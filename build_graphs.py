from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
from PIL import Image
import numpy as np

wb = load_workbook('recdata.xlsx')
ws = wb['Sheet']
# print(wb.sheetnames)
#
# for row in ws.values:
#    for value in row:
#      print(value)

columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O',
           'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
           'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM'
           ]

hoc, foc, doc, toc, soc, noc = 0,0,0,0,0,0 # overall count

res = {}

for xcel_clms in columns[1:26]:
    facename = ws[xcel_clms + str(2)].value
    print(facename)
    res[facename] = {"Frustrated": 0, "Disagree": 0,
                     "Tense": 0, "Happy": 0,
                     "Surprise": 0, "Neutral": 0}

    for frm_no in range(4, 1005):
        emotion = ws[xcel_clms + str(frm_no)].value
        if emotion == "Happy":
            res[facename]["Happy"] += 1
        if emotion == "Frustrated":
            res[facename]["Frustrated"] += 1
        if emotion == "Disagree":
            res[facename]["Disagree"] += 1
        if emotion == "Tense":
            res[facename]["Tense"] += 1
        if emotion == "Surprise":
            res[facename]["Surprise"] += 1
        if emotion == "Neutral":
            res[facename]["Neutral"] += 1

    hc = res[facename]["Happy"]  # happy count
    fc = res[facename]["Frustrated"]
    dc = res[facename]["Disagree"]
    tc = res[facename]["Tense"]
    sc = res[facename]["Surprise"]
    nc = res[facename]["Neutral"]


    hoc += hc
    foc += fc
    doc += dc
    toc += tc
    soc += sc
    noc += nc




cls_left = [1, 2, 3, 4, 5, 6]

total_count = hoc + foc + doc + toc + soc + noc

# Code th plot the graphs using excel file.
axes = plt.gca()
axes.set_ylim([1, 100])
cls_heights = [(hoc / total_count) * 100, (foc / total_count) * 100, (doc / total_count) * 100,
               (toc / total_count) * 100, (soc / total_count) * 100, (noc / total_count) * 100]

tick_label = ['Happy', 'Frustrated', 'Disagree', 'Tense', 'Surprise', 'Neutral']

plt.bar(cls_left, cls_heights, tick_label=tick_label,
        width=0.8, color=(0.0, 0.7, 0.7, 0.5), edgecolor='blue')

plt.xlabel('emotions')

plt.ylabel('percentage')
plt.title( 'class statistics\nTotal Frames: 1000\nTotal Students: 23')

plt.savefig( 'class_statistics.png')
